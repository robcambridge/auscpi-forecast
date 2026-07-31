"""Parse NSW rental bond lodgement workbooks into a mix-controlled rent index.

WHY THIS SERIES. A bond is lodged when a tenancy starts, so every row is a newly
agreed rent on a signed lease. The ABS prices rents across the whole stock of
dwellings, so measured rents only move as sitting leases roll over — roughly a
year. New-lease rents therefore lead measured rents by 6–12 months close to
mechanically, which is the basis of the roll-through model (docs/ROADMAP.md,
Phase 4). Unlike asking rents these are transacted, so there is no listing or
withdrawal bias.

WHY A PLAIN MEDIAN WILL NOT DO, AND THIS IS THE WHOLE DESIGN. The published median
of all lodgements in a month moves whenever the *mix* of what was leased moves. A
month that happens to lease more one-bedroom units than houses prints a lower
median even if not one landlord changed a price. The CPI measures constant-quality
price change, so a series contaminated by composition is not merely noisier — it
is measuring a different quantity, and regressing ABS rents on it would estimate
the roll-through lag from mix drift.

So the index here is fixed-weight. Within each month the median rent is taken per
stratum (dwelling type x bedrooms), each stratum is expressed relative to its own
base-period median, and the strata are recombined on weights held FIXED from the
base window. Composition can then only enter through which strata exist at all,
which the minimum cell rule below bounds. `median_weekly_rent` is carried
alongside precisely so the contamination stays visible: when the two disagree, the
gap is mix.

STRUCTURE VERIFIED, NOT ASSUMED. Checked across all 54 published monthly workbooks
(2022-01 to 2026-06, 1,465,157 rows) rather than inferred from one file:

  - The data is always the FIRST sheet. The second is variously "Definitions",
    "Definition", "Definitions " (trailing space) or "Sheet2", so selecting the
    data sheet by name would break on a quarter of the archive.
  - The header is always the third row, under a title row and a blank one.
  - The columns are always exactly Lodgement Date, Postcode, Dwelling Type,
    Bedrooms, Weekly Rent, and nothing else.
  - Every file holds exactly one lodgement month — 100% of rows in all 54 files,
    so months neither overlap between files nor straggle within one. `file_period`
    still measures this rather than trusting it.

Bedrooms and Weekly Rent arrive as TEXT, not numbers, because both use "U" for
unknown. Coercing them numerically without saying so would turn 1.4% of the
sample into silent NaN.

WHAT IS DISCARDED. Every rule below drops real published rows, so each one is
counted and reported rather than applied quietly — a source that changes shape
shows up as a moved count instead of a moved index:

  - Dwelling type O ("other") and U ("unknown"). The published Definitions sheet
    says Other "may include rented rooms, garages and car spaces", which are not
    dwellings the CPI prices; Unknown cannot be assigned a stratum. Together ~7%.
  - Bedrooms U, and above MAX_BEDROOMS. The published values run to 30, which is a
    boarding house or a typo rather than a dwelling.
  - Weekly rent U, and outside [MIN_WEEKLY_RENT, MAX_WEEKLY_RENT]. About 1,500
    rows in 1.44M sit at $1–50 or $5,000–10,000: car spaces and data entry at one
    end, prestige and probable errors at the other. A median resists them, but
    they are removed before stratification so one row cannot drag a thin cell.

WHAT IS STILL WEAK:

  - NSW only, where the CPI target is national. The same limitation the fuel
    component has, and it needs the same NSW-to-national step later.
  - No geography within NSW. Postcode is carried through to the curated records
    untouched, but Sydney is not split from the rest of the state, because the
    postcode-to-Greater-Sydney correspondence is an ABS product that should be
    read rather than guessed at from postcode ranges. The CPI is a capital-city
    measure, so this split is worth doing before the roll-through regression.
  - Strata are dwelling type x bedrooms only. Within a stratum, quality still
    drifts — a renovated two-bedroom flat and a tired one are one cell.
  - Lodgement date is when the bond reached Fair Trading, not when the lease was
    signed or started. That inserts a short, probably stable administrative lag
    ahead of the economic one the roll-through model estimates.
  - The base window is the earliest months in the build, so index LEVELS are only
    comparable within a single build. See `rent_index`.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, fields
from typing import Any

import openpyxl
import pandas as pd

from auscpi.periods import period_end

#: 0-indexed row carrying the column names, under a title row and a blank one.
HEADER_ROW = 2

EXPECTED_COLUMNS = (
    "Lodgement Date",
    "Postcode",
    "Dwelling Type",
    "Bedrooms",
    "Weekly Rent",
)

#: Published dwelling-type codes, from the workbook's own Definitions sheet.
DWELLING_NAMES = {
    "F": "flat/unit",
    "H": "house",
    "T": "terrace/townhouse/semi-detached",
    "O": "other",
    "U": "unknown",
}

#: The codes that are dwellings the CPI prices. O is explicitly documented as
#: possibly a garage or a car space; U cannot be assigned a stratum.
INDEX_DWELLINGS = ("F", "H", "T")

#: Value both Bedrooms and Weekly Rent use for "not supplied".
UNKNOWN = "U"

MIN_WEEKLY_RENT = 50.0
MAX_WEEKLY_RENT = 5000.0
MAX_BEDROOMS = 5

#: Lodgements a stratum needs in a month before its median is trusted. A median
#: over a handful of leases is noise; the cell is dropped and the fixed weights
#: renormalise over the strata that survive.
MIN_CELL_LODGEMENTS = 30

#: Months of history that set the fixed weights and the per-stratum base level.
BASE_WINDOW_MONTHS = 12

RECORD_COLUMNS = [
    "period",
    "lodgement_date",
    "postcode",
    "dwelling_type",
    "bedrooms",
    "weekly_rent",
]


@dataclass(frozen=True)
class Rejections:
    """Rows dropped by each cleaning rule.

    Reported rather than logged away: these counts are how a silent change in the
    published file — a new code, a units change, a column that starts arriving
    blank — becomes visible as something other than a mysteriously moved index.
    """

    outside_file_month: int = 0
    unknown_dwelling_type: int = 0
    non_dwelling_type: int = 0
    unknown_bedrooms: int = 0
    bedrooms_out_of_range: int = 0
    unknown_rent: int = 0
    rent_out_of_range: int = 0
    unparseable: int = 0

    @property
    def total(self) -> int:
        return sum(getattr(self, f.name) for f in fields(self))

    def __add__(self, other: Rejections) -> Rejections:
        return Rejections(
            **{f.name: getattr(self, f.name) + getattr(other, f.name) for f in fields(self)}
        )

    def as_dict(self) -> dict[str, int]:
        return {f.name: getattr(self, f.name) for f in fields(self)}


def parse_workbook(blob: bytes) -> pd.DataFrame:
    """One published workbook to a frame, as published — no cleaning, no coercion.

    Bedrooms and Weekly Rent stay as text on purpose; they carry "U" for unknown
    and the decision about what to do with that belongs in `clean_records`, where
    it can be counted.
    """
    workbook = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        # By position, never by name: the second sheet is "Definitions",
        # "Definition", "Definitions " or "Sheet2" depending on the month.
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)

        header: tuple[Any, ...] | None = None
        for i, row in enumerate(rows):
            if i == HEADER_ROW:
                header = tuple(str(c).strip() if c is not None else "" for c in row)
                break
        if header is None:
            raise ValueError("workbook has no header row; it is probably not a lodgement file")
        if tuple(header[: len(EXPECTED_COLUMNS)]) != EXPECTED_COLUMNS:
            raise ValueError(
                f"unexpected columns {header!r}; expected {EXPECTED_COLUMNS!r}. The published "
                "layout has changed and the cleaning rules below may no longer mean what they say"
            )

        records = [row[: len(EXPECTED_COLUMNS)] for row in rows if row and row[0] is not None]
    finally:
        workbook.close()

    frame = pd.DataFrame(records, columns=list(EXPECTED_COLUMNS))
    frame["lodgement_date"] = pd.to_datetime(frame["Lodgement Date"], errors="coerce")
    frame["postcode"] = pd.to_numeric(frame["Postcode"], errors="coerce").astype("Int64")
    frame["dwelling_type"] = frame["Dwelling Type"].astype("string").str.strip().str.upper()
    frame["bedrooms_raw"] = frame["Bedrooms"].astype("string").str.strip().str.upper()
    frame["weekly_rent_raw"] = frame["Weekly Rent"].astype("string").str.strip().str.upper()
    return frame.drop(columns=list(EXPECTED_COLUMNS))


def file_period(frame: pd.DataFrame) -> str:
    """The lodgement month a workbook covers, as "YYYY-MM".

    Every published file has so far held exactly one month. Taking the modal month
    rather than assuming it means a file that starts carrying stragglers degrades
    into a counted rejection instead of silently inventing a month with nine rows
    in it.
    """
    months = frame["lodgement_date"].dropna().dt.strftime("%Y-%m")
    if months.empty:
        raise ValueError("workbook has no parseable lodgement dates")
    return str(months.mode().iloc[0])


def clean_records(frame: pd.DataFrame, period: str) -> tuple[pd.DataFrame, Rejections]:
    """Apply the documented exclusions, counting every row each one removes."""
    counts: dict[str, int] = {}
    working = frame.copy()

    before = len(working)
    working = working[working["lodgement_date"].notna()]
    counts["unparseable"] = before - len(working)

    month = working["lodgement_date"].dt.strftime("%Y-%m")
    before = len(working)
    working = working[month == period]
    counts["outside_file_month"] = before - len(working)

    counts["unknown_dwelling_type"] = int((working["dwelling_type"] == UNKNOWN).sum())
    keep_type = working["dwelling_type"].isin(INDEX_DWELLINGS)
    counts["non_dwelling_type"] = int((~keep_type & (working["dwelling_type"] != UNKNOWN)).sum())
    working = working[keep_type]

    counts["unknown_bedrooms"] = int((working["bedrooms_raw"] == UNKNOWN).sum())
    bedrooms = pd.to_numeric(working["bedrooms_raw"], errors="coerce")
    in_range = bedrooms.notna() & (bedrooms >= 0) & (bedrooms <= MAX_BEDROOMS)
    counts["bedrooms_out_of_range"] = int((bedrooms.notna() & ~in_range).sum())
    working = working[in_range]
    bedrooms = bedrooms[in_range]

    counts["unknown_rent"] = int((working["weekly_rent_raw"] == UNKNOWN).sum())
    rent = pd.to_numeric(working["weekly_rent_raw"], errors="coerce")
    priced = rent.notna() & (rent >= MIN_WEEKLY_RENT) & (rent <= MAX_WEEKLY_RENT)
    counts["rent_out_of_range"] = int((rent.notna() & ~priced).sum())
    working = working[priced]

    out = pd.DataFrame(
        {
            "period": period,
            "lodgement_date": working["lodgement_date"].to_numpy(),
            "postcode": working["postcode"].to_numpy(),
            "dwelling_type": working["dwelling_type"].to_numpy(),
            "bedrooms": bedrooms[priced].astype(int).to_numpy(),
            "weekly_rent": rent[priced].to_numpy(),
        }
    )
    return out[RECORD_COLUMNS], Rejections(**counts)


def stratum_medians(records: pd.DataFrame) -> pd.DataFrame:
    """period x (dwelling type, bedrooms) -> median rent and lodgement count.

    Thin cells are dropped here rather than downweighted, because the problem with
    a median over eight leases is not that it is imprecise but that it is
    arbitrary.
    """
    cells = records.groupby(["period", "dwelling_type", "bedrooms"], as_index=False).agg(
        median_rent=("weekly_rent", "median"),
        n=("weekly_rent", "size"),
    )
    return cells[cells["n"] >= MIN_CELL_LODGEMENTS].reset_index(drop=True)


def rent_index(cells: pd.DataFrame, *, base_window: int = BASE_WINDOW_MONTHS) -> pd.DataFrame:
    """Fixed-weight index of stratum medians, so composition cannot move it.

    The weights and the per-stratum base level both come from the EARLIEST
    `base_window` months in the frame, which keeps the construction strictly
    backward-looking: no month's index value depends on data from after it, so a
    series built this way can be fed to a backtest without leaking (CLAUDE.md
    rule 3). Weights taken over the whole sample would be the more conventional
    choice and would quietly break that.

    The consequence to know: index LEVELS are comparable only within one build.
    Rebuild with a longer history and the base window is the same months, but a
    build run `--as-at` an early date has fewer months to draw on and a different
    base. Growth rates are what the model consumes; levels are for reading.

    Each month recombines only the strata it actually has, renormalising the fixed
    weights over them. A stratum appearing or vanishing therefore rescales rather
    than steps the index — the residual composition effect this design cannot
    remove, bounded by MIN_CELL_LODGEMENTS.
    """
    if cells.empty:
        return pd.DataFrame(
            columns=["period", "period_end", "n", "median_weekly_rent", "index", "strata_used"]
        )

    periods = sorted(cells["period"].unique(), key=period_end)
    base_periods = periods[:base_window]
    base = cells[cells["period"].isin(base_periods)]

    stratum = ["dwelling_type", "bedrooms"]
    # Mean of the stratum's monthly medians over the base window, not the median of
    # pooled rents: it weights each base month equally, so an unusually large month
    # cannot set the base level on its own.
    base_level = base.groupby(stratum)["median_rent"].mean()
    base_weight = base.groupby(stratum)["n"].sum()
    base_weight = base_weight / base_weight.sum()

    rows: list[dict[str, Any]] = []
    for period in periods:
        month = cells[cells["period"] == period].set_index(stratum)
        usable = month.index.intersection(base_level.index)
        if usable.empty:
            continue
        weights = base_weight.loc[usable]
        relative = month.loc[usable, "median_rent"] / base_level.loc[usable]
        rows.append(
            {
                "period": period,
                "period_end": period_end(period),
                "n": int(month.loc[usable, "n"].sum()),
                "index": float((weights * relative).sum() / weights.sum() * 100.0),
                "strata_used": len(usable),
            }
        )

    frame = pd.DataFrame(rows)
    frame["index_mom_pct"] = frame["index"].pct_change() * 100.0
    frame["index_yoy_pct"] = frame["index"].pct_change(12) * 100.0
    return frame


def index_frame(records: pd.DataFrame, *, base_window: int = BASE_WINDOW_MONTHS) -> pd.DataFrame:
    """The published monthly series: the mix-controlled index and its raw foil.

    `median_weekly_rent` is the plain median of every cleaned lodgement, which is
    roughly what NSW Fair Trading publishes and is NOT the modelling input. It sits
    here so the mix contamination the index removes can be seen rather than taken
    on trust: where `median_mom_pct` and `index_mom_pct` disagree, the difference
    is composition.
    """
    cells = stratum_medians(records)
    frame = rent_index(cells, base_window=base_window)
    if frame.empty:
        return frame

    raw = records.groupby("period", as_index=False).agg(
        median_weekly_rent=("weekly_rent", "median"),
        n_lodgements=("weekly_rent", "size"),
    )
    frame = frame.merge(raw, on="period", how="left")
    frame["median_mom_pct"] = frame["median_weekly_rent"].pct_change() * 100.0

    return frame[
        [
            "period",
            "period_end",
            "n_lodgements",
            "median_weekly_rent",
            "median_mom_pct",
            "index",
            "index_mom_pct",
            "index_yoy_pct",
            "strata_used",
        ]
    ]
