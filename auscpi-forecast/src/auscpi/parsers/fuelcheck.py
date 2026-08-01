"""Parse NSW FuelCheck price history into a daily and monthly price series.

WHY FUEL IS WORTH THE TROUBLE. `auscpi leverage` ranks automotive fuel third of 87
expenditure classes by how much headline movement it can produce (0.199pp), and it is
the highest-leverage class that is genuinely anticipable rather than merely volatile.
March 2026 makes the case on its own: the ABS index rose 32.79% in one month, which
at a 3.347% basket weight is 1.10pp on the headline from a single class. NSW publishes
every prescribed price at every station, so this component is measured rather than
modelled — the remaining error is the NSW-to-national gap, not the price.

THE ROWS ARE PRICE CHANGES, NOT OBSERVATIONS, AND THIS IS THE WHOLE DESIGN. Each row
is a station posting a new price for one fuel at one moment. A station that holds its
price for a week appears once; a station in the middle of a discount cycle appears
daily. Averaging the Price column therefore weights stations by how often they
re-price, which in a market with a pronounced discount cycle is a direct bias.
Measured against the correct construction it runs about 2.7 c/L low, consistently:

    2026-03   naive mean 226.98   step-function mean 229.82   (-2.85)
    2026-06   naive mean 165.77   step-function mean 168.43   (-2.66)

The sign is not random. Stations shave prices in small decrements through the
discount phase and restore them in one jump, so change events over-sample the cheap
end of the cycle. On a series whose monthly moves are a few per cent, 1.5% of bias
is not a rounding detail.

So a price is treated as a step function: whatever a station last posted is what it
charges until it posts again. `daily_prices` reconstructs that, averages ACROSS
stations for each day, and only then averages days into a month — the order matters,
because averaging events first bakes the bias in.

CARRY-IN IS REQUIRED, AND A MONTH CANNOT BE PARSED ALONE. Forward-fill cannot fill
backwards, so on the first day of a file only stations that happened to post that day
have a price. In March 2026 that was 205 of 2,041 stations — 10% — and they are a
biased 10%, because a station posts precisely when it has just changed its price.
Seeding each station with its last known price from the previous file lifts day-one
coverage to 78%. `build.build_fuelcheck` therefore walks months in order and threads
the carry-in forward; `daily_prices` will accept one but does not fetch it, so the
chaining stays visible in the caller rather than hidden here.

Coverage is reported per day rather than assumed, because the remaining 22% is real:
a station that has not changed price for two months is genuinely unknown until it
does, and a day assembled from three quarters of the network should be legible as
such.

WHAT IS STILL WEAK:

  - NSW only, against a national CPI target. The same limitation the rent component
    has, and it needs the same NSW-to-national step.
  - Unweighted across stations, and there is evidence this matters. A price index
    should weight by volume sold; a quiet rural station currently counts as much as a
    busy metropolitan one. The tell is that IMPROVING station coverage makes the
    series correlate slightly WORSE with the ABS class (0.9840 with no carry-in and
    218 thin days, 0.9727 with a bounded carry and 25). Stations that re-price
    constantly are the busy competitive sites where most fuel is sold, so a crude
    event-weighted sample accidentally approximates volume weighting. Volume data is
    not published, so a proxy would have to be built; postcode is carried through to
    the events so a population weighting could be tried.
  - Unweighted across fuels. The ABS class covers petrol, diesel and LPG on
    expenditure shares. Per-fuel series are produced and the blend is left to the
    caller rather than guessed at here.
  - A price posted and never corrected stays in the step function forever. There is
    no way to distinguish a station that closed from one that simply held its price.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

import pandas as pd

#: Columns as published, in both the xlsx and csv variants of the archive.
EXPECTED_COLUMNS = (
    "ServiceStationName",
    "Address",
    "Suburb",
    "Postcode",
    "Brand",
    "FuelCode",
    "PriceUpdatedDate",
    "Price",
)

#: Fuel codes seen across the archive, most to least common. PDL and DL are premium
#: and regular diesel; E10 and E85 are ethanol blends; B20 is a biodiesel blend.
FUEL_CODES = ("U91", "P95", "P98", "E10", "E85", "DL", "PDL", "LPG", "B20")

#: The code that carries the headline "petrol price" in Australian usage, and the
#: single best match for the unleaded part of the ABS class.
HEADLINE_FUEL = "U91"

#: Cents per litre. Wide enough to survive any plausible price, tight enough to drop
#: a decimal-point error or a placeholder.
MIN_PRICE_CENTS = 50.0
MAX_PRICE_CENTS = 400.0


@dataclass(frozen=True)
class Rejections:
    """Rows dropped by each cleaning rule, reported rather than logged away."""

    unparseable_timestamp: int = 0
    unparseable_price: int = 0
    price_out_of_range: int = 0
    unknown_fuel_code: int = 0

    @property
    def total(self) -> int:
        return (
            self.unparseable_timestamp
            + self.unparseable_price
            + self.price_out_of_range
            + self.unknown_fuel_code
        )

    def as_dict(self) -> dict[str, int]:
        return {
            "unparseable_timestamp": self.unparseable_timestamp,
            "unparseable_price": self.unparseable_price,
            "price_out_of_range": self.price_out_of_range,
            "unknown_fuel_code": self.unknown_fuel_code,
        }

    def __add__(self, other: Rejections) -> Rejections:
        return Rejections(
            self.unparseable_timestamp + other.unparseable_timestamp,
            self.unparseable_price + other.unparseable_price,
            self.price_out_of_range + other.price_out_of_range,
            self.unknown_fuel_code + other.unknown_fuel_code,
        )


#: Column whose presence identifies the header row.
HEADER_MARKER = "ServiceStationName"

#: How far into a file to look for the header before giving up. The archive is
#: inconsistent: csv files put the header on the first line and so do most xlsx
#: files, but ten of them carry a "Price History Checks" title row and a blank one
#: first, so the header sits on the third. Assuming position parses two thirds of the
#: archive and fails loudly on the rest, which is how this was found.
HEADER_SEARCH_ROWS = 8


def parse_price_events(blob: bytes) -> pd.DataFrame:
    """One published price-history file to a frame, as published.

    The archive ships the same table as xlsx for most months and csv for others, so
    the format is detected from the payload rather than trusted from the resource
    metadata — a mislabelled resource should not decide how bytes are read. xlsx is
    a zip and starts with PK; the csv carries a UTF-8 BOM.

    The header is located rather than assumed; see HEADER_SEARCH_ROWS.
    """
    if blob[:2] == b"PK":
        import openpyxl

        workbook = openpyxl.load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        try:
            sheet = workbook[workbook.sheetnames[0]]
            rows = sheet.iter_rows(values_only=True)
            header: list[str] | None = None
            for i, row in enumerate(rows):
                cells = [str(c).strip() if c is not None else "" for c in row]
                if HEADER_MARKER in cells:
                    header = cells
                    break
                if i >= HEADER_SEARCH_ROWS:
                    break
            if header is None:
                raise ValueError(
                    f"no header row containing {HEADER_MARKER!r} in the first "
                    f"{HEADER_SEARCH_ROWS} rows; this is probably not a price-history file"
                )
            # The iterator stopped ON the header, so what remains is data.
            frame = pd.DataFrame(list(rows), columns=header)
        finally:
            workbook.close()
    else:
        peek = blob[:8192].decode("utf-8-sig", "replace").splitlines()
        skip = next(
            (i for i, line in enumerate(peek[:HEADER_SEARCH_ROWS]) if HEADER_MARKER in line), None
        )
        if skip is None:
            raise ValueError(
                f"no header row containing {HEADER_MARKER!r} in the first "
                f"{HEADER_SEARCH_ROWS} lines; this is probably not a price-history file"
            )
        frame = pd.read_csv(
            io.BytesIO(blob), encoding="utf-8-sig", low_memory=False, skiprows=skip
        )

    frame.columns = [str(c).strip() for c in frame.columns]
    missing = [c for c in EXPECTED_COLUMNS if c not in frame.columns]
    if missing:
        raise ValueError(
            f"price history is missing {missing}; got {list(frame.columns)}. The "
            "published layout has changed and the cleaning rules no longer mean what "
            "they say"
        )
    return frame[list(EXPECTED_COLUMNS)]


def _to_timestamp(values: pd.Series) -> pd.Series:
    """Parse PriceUpdatedDate, which is not one format across the archive.

    THIS IS A SILENT-CORRUPTION GUARD, NOT A CONVENIENCE. Most files carry ISO
    timestamps ("2026-03-01 00:05:26") but three of the 42 captured use Australian
    day-first slashes ("13/09/2023 12:04:31 AM"). Left to infer, pandas reads the
    opening rows of such a file as MONTH-first, silently mis-dates every row whose
    day is 12 or lower — 01/09/2023 becomes 9 January — and only fails outright on
    days 13 and up. The failure is therefore visible as ~70% unparseable while the
    other 30% is quietly wrong, which is the worse half.

    So ISO is tried strictly first, and only what it rejects is retried day-first.
    Never let this fall back to inference.
    """
    if pd.api.types.is_datetime64_any_dtype(values):
        return values
    text = values.astype("string")
    parsed = pd.to_datetime(text, errors="coerce", format="ISO8601")
    missing = parsed.isna() & text.notna()
    if missing.any():
        parsed = parsed.fillna(
            pd.to_datetime(text.where(missing), errors="coerce", dayfirst=True)
        )
    return parsed


def clean_events(frame: pd.DataFrame) -> tuple[pd.DataFrame, Rejections]:
    """Coerce types and drop what cannot be priced, counting each rule."""
    working = frame.copy()
    timestamp = _to_timestamp(working["PriceUpdatedDate"])
    price = pd.to_numeric(working["Price"], errors="coerce")
    fuel = working["FuelCode"].astype("string").str.strip().str.upper()

    bad_time = int(timestamp.isna().sum())
    bad_price = int(price.isna().sum())
    in_range = price.between(MIN_PRICE_CENTS, MAX_PRICE_CENTS)
    out_of_range = int((price.notna() & ~in_range).sum())
    known_fuel = fuel.isin(FUEL_CODES)
    unknown_fuel = int((~known_fuel).sum())

    keep = timestamp.notna() & in_range & known_fuel
    out = pd.DataFrame(
        {
            # Name alone is not unique — chains repeat it across sites — so the
            # address is part of the key.
            "station": (
                working["ServiceStationName"].astype("string").fillna("")
                + "|"
                + working["Address"].astype("string").fillna("")
            )[keep].to_numpy(),
            "postcode": working["Postcode"][keep].to_numpy(),
            "brand": working["Brand"][keep].to_numpy(),
            "fuel": fuel[keep].to_numpy(),
            "ts": timestamp[keep].to_numpy(),
            "price": price[keep].to_numpy(),
        }
    )
    return out, Rejections(bad_time, bad_price, out_of_range, unknown_fuel)


def daily_prices(
    events: pd.DataFrame,
    *,
    fuel: str = HEADLINE_FUEL,
    carry_in: pd.Series | None = None,
) -> pd.DataFrame:
    """Mean posted price across stations, per day, from the step function.

    `carry_in` maps station to its last known price before this file. Supply it or
    the first days are assembled from the small, biased subset that happened to post
    early — see the module docstring. Returns one row per day with the mean and the
    number of stations behind it, so a thin day is visible rather than silent.
    """
    subset = events[events["fuel"] == fuel]
    if subset.empty:
        return pd.DataFrame(columns=["date", "fuel", "mean_price", "stations"])

    day = subset["ts"].dt.floor("D")
    last = (
        subset.assign(day=day)
        .sort_values("ts")
        .groupby(["station", "day"])["price"]
        .last()
        .unstack("day")
        .sort_index(axis=1)
    )
    if carry_in is not None and not carry_in.empty:
        # The index must be the UNION, not this file's stations. A station that held
        # its price for the whole month posts nothing and so appears nowhere in the
        # events — which is precisely the case carry-in exists to cover. Reindexing
        # onto the events alone silently drops exactly the stations being rescued.
        combined = last.index.union(carry_in.index)
        last = last.reindex(combined)
        # Seeded before the first day so the forward fill has somewhere to start.
        last.insert(0, "_carry_in", carry_in.reindex(combined))
    filled = last.ffill(axis=1)
    if "_carry_in" in filled.columns:
        filled = filled.drop(columns="_carry_in")

    return pd.DataFrame(
        {
            "date": [d.date() for d in filled.columns],
            "fuel": fuel,
            "mean_price": filled.mean(axis=0).to_numpy(),
            "stations": filled.notna().sum(axis=0).to_numpy(),
        }
    )


def closing_prices(events: pd.DataFrame, *, fuel: str = HEADLINE_FUEL) -> pd.Series:
    """Each station's last posted price in this file, to carry into the next one."""
    subset = events[events["fuel"] == fuel]
    if subset.empty:
        return pd.Series(dtype="float64")
    return subset.sort_values("ts").groupby("station")["price"].last()


def monthly_prices(daily: pd.DataFrame) -> pd.DataFrame:
    """Monthly mean of the daily cross-sectional means, plus coverage.

    Averaging days rather than events is the point: every day counts once regardless
    of how many stations happened to re-price on it, so a volatile week cannot
    outvote a quiet one.
    """
    if daily.empty:
        return pd.DataFrame(
            columns=["period", "fuel", "mean_price", "days", "min_stations", "mean_stations"]
        )
    frame = daily.copy()
    frame["period"] = pd.to_datetime(frame["date"]).dt.strftime("%Y-%m")
    grouped = frame.groupby(["period", "fuel"], as_index=False).agg(
        mean_price=("mean_price", "mean"),
        days=("mean_price", "size"),
        min_stations=("stations", "min"),
        mean_stations=("stations", "mean"),
    )
    grouped["mom_pct"] = grouped.groupby("fuel")["mean_price"].pct_change() * 100.0
    return grouped
